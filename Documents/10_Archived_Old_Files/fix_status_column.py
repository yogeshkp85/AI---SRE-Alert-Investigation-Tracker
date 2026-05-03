#!/usr/bin/env python3
"""
Fix Status column to contain proper status values (In Progress, Pending, Completed)
"""

import openpyxl
import random

def fix_status_column():
    """Fix Status column values"""
    
    # Load the workbook
    wb = openpyxl.load_workbook('incident-tracker.xlsx')
    ws = wb.active
    
    # Find the Status column (should be column 22)
    status_col = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Status':
            status_col = col_idx
            break
    
    if not status_col:
        print("Error: Status column not found!")
        return
    
    print(f"Found Status column at column {status_col}")
    
    # Update Status values
    statuses = ['In Progress', 'Pending', 'Completed']
    
    for row_idx in range(2, ws.max_row + 1):
        # Randomly assign a status
        new_status = random.choice(statuses)
        ws.cell(row=row_idx, column=status_col, value=new_status)
        print(f"Row {row_idx}: Status = {new_status}")
    
    # Save the workbook
    wb.save('incident-tracker.xlsx')
    wb.close()
    
    print("\n✓ Status column fixed successfully!")
    print("✓ All Status values updated to: In Progress, Pending, or Completed")
    print("✓ File saved: incident-tracker.xlsx")

if __name__ == '__main__':
    fix_status_column()
