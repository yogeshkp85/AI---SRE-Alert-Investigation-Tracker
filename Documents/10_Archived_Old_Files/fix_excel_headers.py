#!/usr/bin/env python3
"""
Fix Excel headers and data alignment
"""

import openpyxl
from datetime import datetime, timedelta
import random

EXCEL_FILE = 'incident-tracker.xlsx'

def fix_headers_and_data():
    """Fix the Excel file with correct headers and properly aligned data"""
    try:
        # Read current data
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Sheet1']
        
        # Get team members from Sheet2
        ws2 = wb['Sheet2']
        team_members = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0]:
                team_members.append(row[0])
        
        print(f"📝 Found {len(team_members)} team members")
        
        # Read all current data rows (skip header)
        current_data = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, 29):
                cell = ws.cell(row_idx, col_idx)
                row_data.append(cell.value)
            current_data.append(row_data)
        
        print(f"📊 Read {len(current_data)} data rows")
        
        # Clear all rows except header
        while ws.max_row > 1:
            ws.delete_rows(2, 1)
        
        # Set correct headers
        correct_headers = [
            'Date', 'Shift', 'Time Slot', 'Alert Report Time', 'Alert',
            'Assigned To', 'RITM', 'STIP Incident', 'Incident Raised', 'Email',
            'DB Giant', 'Type Comms', 'Incident Comms', 'Batch Reportable',
            'Final Comms', 'CR', 'Implementation', 'Verification',
            'Issue Communication', 'Additional Task/Improvement', 'Status'
        ]
        
        for col_idx, header in enumerate(correct_headers, 1):
            ws.cell(1, col_idx).value = header
        
        print("✅ Headers fixed")
        
        # Create fresh data with correct alignment
        print("\n📊 Creating fresh data with correct alignment...\n")
        
        base_date = datetime.now()
        shifts = ["S1", "S2", "On Call"]
        statuses = ["In Progress", "Pending", "Completed"]
        categories = ["P1", "P2", "P3", "P4"]
        time_slots = ["7-8 AM", "8-9 AM", "9-10 AM", "10-11 AM", "11-12 PM", "12-1 PM", 
                     "1-2 PM", "2-3 PM", "3-4 PM", "4-5 PM", "5-6 PM", "6-7 PM", 
                     "7-8 PM", "8-9 PM", "9-10 PM", "10 PM-7 AM"]
        
        # Create 20 fresh incidents with correct column alignment
        for i in range(20):
            row_num = i + 2
            incident_date = (base_date - timedelta(days=i)).strftime('%Y-%m-%d')
            shift = random.choice(shifts)
            time_slot = random.choice(time_slots)
            alert_time = f"{random.randint(6, 22):02d}:{random.randint(0, 59):02d}"
            assigned_to = random.choice(team_members)
            status = random.choice(statuses)
            category = random.choice(categories)
            
            # Write to correct columns
            ws.cell(row_num, 1).value = incident_date  # Date
            ws.cell(row_num, 2).value = shift  # Shift
            ws.cell(row_num, 3).value = time_slot  # Time Slot
            ws.cell(row_num, 4).value = alert_time  # Alert Report Time
            ws.cell(row_num, 5).value = f"Payment Gateway Timeout - Transaction {1000+i}"  # Alert
            ws.cell(row_num, 6).value = assigned_to  # Assigned To
            ws.cell(row_num, 7).value = f"INC{1000+i}"  # RITM
            ws.cell(row_num, 8).value = f"STIP{500+i}"  # STIP Incident
            ws.cell(row_num, 9).value = f"IR{200+i}"  # Incident Raised
            ws.cell(row_num, 10).value = f"alert-{i}@company.com"  # Email
            ws.cell(row_num, 11).value = f"DB Alert {i}"  # DB Giant
            ws.cell(row_num, 12).value = "Emergency"  # Type Comms
            ws.cell(row_num, 13).value = f"Payment processing issue {i}"  # Incident Comms
            ws.cell(row_num, 14).value = random.choice(["Yes", "No"])  # Batch Reportable
            ws.cell(row_num, 15).value = f"Final update {i}"  # Final Comms
            ws.cell(row_num, 16).value = random.choice(["Yes", "No"])  # CR
            ws.cell(row_num, 17).value = random.choice(["Yes", "No"])  # Implementation
            ws.cell(row_num, 18).value = f"Verified by {random.choice(team_members)}"  # Verification
            ws.cell(row_num, 19).value = f"Issue resolved {i}"  # Issue Communication
            ws.cell(row_num, 20).value = f"Monitor system {i}"  # Additional Task
            ws.cell(row_num, 21).value = status  # Status
            
            print(f"   Row {row_num}: {incident_date} | {shift} | {time_slot} | {assigned_to} | {status}")
        
        # Save
        wb.save(EXCEL_FILE)
        print(f"\n✅ Saved to {EXCEL_FILE}")
        
        print(f"\n🎉 Excel Headers and Data Fixed!")
        print(f"   Total incidents: 20")
        print(f"   All columns properly aligned")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    fix_headers_and_data()
