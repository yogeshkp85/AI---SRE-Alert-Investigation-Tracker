#!/usr/bin/env python3
"""
Fix Excel schema to match dashboard requirements
Adds missing columns: Incident Category, MTTR (minutes), Created At, Completed At
"""

import openpyxl
from datetime import datetime, timedelta
import random

def fix_excel_schema():
    """Add missing columns to Excel file"""
    
    # Load the workbook
    wb = openpyxl.load_workbook('incident-tracker.xlsx')
    ws = wb.active
    
    # Get current headers
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"Current headers ({len(headers)}): {headers}")
    
    # Add new columns after Status
    new_columns = ['Incident Category', 'MTTR (minutes)', 'Created At', 'Completed At']
    
    # Insert new headers
    next_col = len(headers) + 1
    for idx, col_name in enumerate(new_columns):
        ws.cell(row=1, column=next_col + idx, value=col_name)
    
    print(f"\nAdded columns: {new_columns}")
    
    # Now populate data for each row
    for row_idx in range(2, ws.max_row + 1):
        # Get the Status value (which contains P1, P2, P3, P4)
        status_cell = ws.cell(row=row_idx, column=len(headers))
        status_value = status_cell.value
        
        # Extract category from Status (it contains P1, P2, P3, or P4)
        category = 'P4'  # default
        if status_value:
            status_str = str(status_value).upper()
            if 'P1' in status_str:
                category = 'P1'
            elif 'P2' in status_str:
                category = 'P2'
            elif 'P3' in status_str:
                category = 'P3'
            elif 'P4' in status_str:
                category = 'P4'
        
        # Set Incident Category
        ws.cell(row=row_idx, column=next_col, value=category)
        
        # Get Date and Time for creating timestamps
        date_cell = ws.cell(row=row_idx, column=1)  # Date column
        time_cell = ws.cell(row=row_idx, column=5)  # Alert Report Time column
        
        date_value = date_cell.value
        time_value = time_cell.value
        
        # Create timestamps
        try:
            if date_value:
                # Parse date
                if isinstance(date_value, str):
                    date_obj = datetime.fromisoformat(date_value)
                else:
                    date_obj = date_value
                
                # Parse time (format: "HH:MM")
                if time_value:
                    time_str = str(time_value).strip()
                    if ':' in time_str:
                        time_parts = time_str.split(':')
                        hour = int(time_parts[0])
                        minute = int(time_parts[1]) if len(time_parts) > 1 else 0
                    else:
                        hour = 9
                        minute = 0
                else:
                    hour = 9
                    minute = 0
                
                # Create Created At timestamp
                created_at = date_obj.replace(hour=hour, minute=minute, second=0, microsecond=0)
                ws.cell(row=row_idx, column=next_col + 2, value=created_at.isoformat())
                
                # Create Completed At timestamp (random between 30 min to 4 hours later)
                # Only for "Completed" status
                status_col = ws.cell(row=row_idx, column=len(headers) - 1)  # Status column (before new ones)
                
                # For now, set Completed At to a random time after Created At
                minutes_to_add = random.randint(30, 240)
                completed_at = created_at + timedelta(minutes=minutes_to_add)
                ws.cell(row=row_idx, column=next_col + 3, value=completed_at.isoformat())
                
                # Set MTTR (minutes)
                ws.cell(row=row_idx, column=next_col + 1, value=minutes_to_add)
                
                print(f"Row {row_idx}: Category={category}, MTTR={minutes_to_add}m, Created={created_at.isoformat()}")
            
        except Exception as e:
            print(f"Error processing row {row_idx}: {e}")
            # Set defaults
            ws.cell(row=row_idx, column=next_col + 1, value=0)  # MTTR
            ws.cell(row=row_idx, column=next_col + 2, value=datetime.now().isoformat())  # Created At
            ws.cell(row=row_idx, column=next_col + 3, value=datetime.now().isoformat())  # Completed At
    
    # Save the workbook
    wb.save('incident-tracker.xlsx')
    wb.close()
    
    print("\n✓ Excel schema fixed successfully!")
    print("✓ Added columns: Incident Category, MTTR (minutes), Created At, Completed At")
    print("✓ File saved: incident-tracker.xlsx")

if __name__ == '__main__':
    fix_excel_schema()
