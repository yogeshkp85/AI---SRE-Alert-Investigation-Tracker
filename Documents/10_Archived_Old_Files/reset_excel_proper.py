#!/usr/bin/env python3
"""
Reset Excel file with proper structure:
- Sheet1: Incidents (with all required columns)
- Sheet2: Team Members (already exists)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# Team members list
TEAM_MEMBERS = [
    'Nilam Patel', 'Shital Waghmare', 'Dnyaneshwar Chaudhary', 'Gunjan Pujara',
    'Amal P Raj', 'Midhun Pushparaj', 'Aparna KS', 'Rhutuja Aher',
    'Shweta Patil', 'Prasad Khopade', 'Hitesh Pitrubhakta', 'Navjyot Bhosale',
    'Deepak Sahoo', 'Shubham Shrivastava', 'Riyaz Husain', 'Vertika Singh',
    'Hunny Kumar'
]

SHIFTS = ['S1', 'S2', 'On Call']
CATEGORIES = ['P1', 'P2', 'P3', 'P4']
STATUSES = ['In Progress', 'Pending', 'Completed']
TIME_SLOTS = ['7-8 AM', '8-9 AM', '9-10 AM', '10-11 AM', '11-12 PM', '12-1 PM', 
              '1-2 PM', '2-3 PM', '3-4 PM', '4-5 PM', '5-6 PM', '6-7 PM', 
              '7-8 PM', '8-9 PM', '9-10 PM', '10 PM-7 AM']

# Create workbook
wb = openpyxl.Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Create Sheet1 for Incidents
ws1 = wb.create_sheet('Sheet1', 0)

# Define headers for incidents
headers = [
    'Date', 'Shift', 'Incident Category', 'Shift Lead', 'Time Slot', 'Alert Report Time',
    'Alert', 'Assigned To', 'Status', 'RITM', 'STIP Incident', 'Incident Raised',
    'Email', 'DB Giant', 'Type Comms', 'Incident Comms', 'Batch Reportable',
    'Final Comms', 'CR', 'Implementation', 'Verification', 'Issue Communication',
    'Additional Task/Improvement', 'Created At', 'Completed At', 'MTTR (minutes)',
    'Last Modified By', 'Last Modified At'
]

# Write headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Generate 25 sample incidents
base_date = datetime.now() - timedelta(days=30)

for row_idx in range(2, 27):  # 25 incidents
    incident_num = row_idx - 1
    
    # Generate data
    incident_date = base_date + timedelta(days=random.randint(0, 30))
    shift = random.choice(SHIFTS)
    category = random.choice(CATEGORIES)
    shift_lead = random.choice(TEAM_MEMBERS)
    time_slot = random.choice(TIME_SLOTS)
    alert_time = f"{random.randint(6, 22):02d}:{random.randint(0, 59):02d}"
    assigned_to = random.choice(TEAM_MEMBERS)
    status = random.choice(STATUSES)
    
    # Create alert description
    alert_desc = f"Payment transaction incident #{incident_num}. Alert triggered for transaction processing delay."
    
    # Calculate MTTR if completed
    mttr = None
    completed_at = None
    if status == 'Completed':
        completed_at = incident_date + timedelta(hours=random.randint(1, 8))
        mttr = int((completed_at - incident_date).total_seconds() / 60)
    
    # Write data
    data = [
        incident_date.strftime('%Y-%m-%d'),  # Date
        shift,  # Shift
        category,  # Incident Category
        shift_lead,  # Shift Lead
        time_slot,  # Time Slot
        alert_time,  # Alert Report Time
        alert_desc,  # Alert
        assigned_to,  # Assigned To
        status,  # Status
        f"RITM{incident_num:06d}",  # RITM
        f"STIP{incident_num:06d}",  # STIP Incident
        f"INC{incident_num:06d}",  # Incident Raised
        f"Alert for incident {incident_num}",  # Email
        f"DB Issue {incident_num}",  # DB Giant
        "Email",  # Type Comms
        f"Communication for incident {incident_num}",  # Incident Comms
        random.choice(['Yes', 'No']),  # Batch Reportable
        f"Final communication {incident_num}",  # Final Comms
        random.choice(['Yes', 'No']),  # CR
        random.choice(['Yes', 'No']),  # Implementation
        f"Verification details for incident {incident_num}",  # Verification
        f"Issue communication for incident {incident_num}",  # Issue Communication
        f"Additional tasks and improvements for incident {incident_num}",  # Additional Task
        incident_date.isoformat(),  # Created At
        completed_at.isoformat() if completed_at else None,  # Completed At
        mttr,  # MTTR (minutes)
        "System",  # Last Modified By
        datetime.now().isoformat()  # Last Modified At
    ]
    
    for col_idx, value in enumerate(data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Adjust column widths
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 30
ws1.column_dimensions['H'].width = 15
ws1.column_dimensions['I'].width = 12
for col in range(10, 28):
    ws1.column_dimensions[chr(64 + col)].width = 15

# Create/Update Sheet2 for Team Members
if 'Sheet2' in wb.sheetnames:
    ws2 = wb['Sheet2']
    ws2.delete_rows(1, ws2.max_row)
else:
    ws2 = wb.create_sheet('Sheet2', 1)

# Write team member headers
team_headers = ['Name', 'Email', 'Phone']
for col_idx, header in enumerate(team_headers, start=1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")

# Write team members
team_data = [
    ('Nilam Patel', 'nilam.patel@dbindia.com', '9892542648'),
    ('Shital Waghmare', 'Shital@dbindia.com', '9876543211'),
    ('Dnyaneshwar Chaudhary', 'dnyaneshwar@dbindia.com', '9876543212'),
    ('Gunjan Pujara', 'gunjan@dbindia.com', '9876543213'),
    ('Amal P Raj', 'amal@dbindia.com', '9876543214'),
    ('Midhun Pushparaj', 'midhun@dbindia.com', '9876543215'),
    ('Aparna KS', 'aparna@dbindia.com', '9876543216'),
    ('Rhutuja Aher', 'rhutuja@dbindia.com', '9876543217'),
    ('Shweta Patil', 'shweta@dbindia.com', '9876543218'),
    ('Prasad Khopade', 'prasad@dbindia.com', '9876543219'),
    ('Hitesh Pitrubhakta', 'hitesh@dbindia.com', '9876543220'),
    ('Navjyot Bhosale', 'navjyot@dbindia.com', '9876543221'),
    ('Deepak Sahoo', 'deepak@dbindia.com', '9876543222'),
    ('Shubham Shrivastava', 'shubham@dbindia.com', '9876543223'),
    ('Riyaz Husain', 'riyaz@dbindia.com', '9876543224'),
    ('Vertika Singh', 'vertika@dbindia.com', '9876543225'),
    ('Hunny Kumar', 'hunny@dbindia.com', '9876543226'),
]

for row_idx, (name, email, phone) in enumerate(team_data, start=2):
    ws2.cell(row=row_idx, column=1).value = name
    ws2.cell(row=row_idx, column=2).value = email
    ws2.cell(row=row_idx, column=3).value = phone

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 15

# Save workbook
wb.save('incident-tracker.xlsx')
print("✓ Excel file reset successfully!")
print(f"  - Sheet1: 25 incidents with all required columns")
print(f"  - Sheet2: 17 team members")
print(f"  - Total columns: {len(headers)}")
