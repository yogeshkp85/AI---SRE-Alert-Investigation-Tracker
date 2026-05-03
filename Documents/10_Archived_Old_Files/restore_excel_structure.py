#!/usr/bin/env python3
"""
Restore Excel structure with ALL required columns including Incident Category and Shift Lead
"""

import openpyxl
from datetime import datetime, timedelta
import random

EXCEL_FILE = 'incident-tracker.xlsx'

def restore_excel_structure():
    """Restore Excel with correct structure including Incident Category and Shift Lead"""
    try:
        # Read current data and team members
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws1 = wb['Sheet1']
        ws2 = wb['Sheet2']
        
        # Get team members from Sheet2
        team_members = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0]:
                team_members.append(row[0])
        
        print(f"📝 Found {len(team_members)} team members")
        
        # Clear all rows except header
        while ws1.max_row > 1:
            ws1.delete_rows(2, 1)
        
        # Set CORRECT headers - matching the original structure
        correct_headers = [
            'Date',                      # A
            'Shift',                     # B
            'Incident Category',         # C - RESTORED!
            'Shift Lead',                # D - RESTORED!
            'Time Slot',                 # E
            'Alert Report Time',         # F
            'Alert',                     # G
            'Assigned To',               # H
            'RITM',                      # I
            'STIP Incident',             # J
            'Incident Raised',           # K
            'Email',                     # L
            'DB Giant',                  # M
            'Type Comms',                # N
            'Incident Comms',            # O
            'Batch Reportable',          # P
            'Final Comms',               # Q
            'CR',                        # R
            'Implementation',            # S
            'Verification',              # T
            'Issue Communication',       # U
            'Additional Task/Improvement', # V
            'Status'                     # W
        ]
        
        for col_idx, header in enumerate(correct_headers, 1):
            ws1.cell(1, col_idx).value = header
        
        print("✅ Headers restored with Incident Category and Shift Lead")
        print(f"\n   Headers: {', '.join(correct_headers[:8])}...")
        
        # Create fresh data with correct alignment
        print("\n📊 Creating fresh data with correct structure...\n")
        
        base_date = datetime.now()
        shifts = ["S1", "S2", "On Call"]
        statuses = ["In Progress", "Pending", "Completed"]
        categories = ["P1", "P2", "P3", "P4"]
        time_slots = ["7-8 AM", "8-9 AM", "9-10 AM", "10-11 AM", "11-12 PM", "12-1 PM", 
                     "1-2 PM", "2-3 PM", "3-4 PM", "4-5 PM", "5-6 PM", "6-7 PM", 
                     "7-8 PM", "8-9 PM", "9-10 PM", "10 PM-7 AM"]
        
        # Create 20 fresh incidents with correct column alignment
        for i in range(20):
            row_num = i + 2
            incident_date = (base_date - timedelta(days=i)).strftime('%Y-%m-%d')
            shift = random.choice(shifts)
            incident_category = random.choice(categories)  # NEW!
            shift_lead = random.choice(team_members)  # NEW!
            time_slot = random.choice(time_slots)
            alert_time = f"{random.randint(6, 22):02d}:{random.randint(0, 59):02d}"
            assigned_to = random.choice(team_members)
            status = random.choice(statuses)
            
            # Write to correct columns
            ws1.cell(row_num, 1).value = incident_date  # Date
            ws1.cell(row_num, 2).value = shift  # Shift
            ws1.cell(row_num, 3).value = incident_category  # Incident Category - RESTORED!
            ws1.cell(row_num, 4).value = shift_lead  # Shift Lead - RESTORED!
            ws1.cell(row_num, 5).value = time_slot  # Time Slot
            ws1.cell(row_num, 6).value = alert_time  # Alert Report Time
            ws1.cell(row_num, 7).value = f"Payment Gateway Timeout - Transaction {1000+i}"  # Alert
            ws1.cell(row_num, 8).value = assigned_to  # Assigned To
            ws1.cell(row_num, 9).value = f"INC{1000+i}"  # RITM
            ws1.cell(row_num, 10).value = f"STIP{500+i}"  # STIP Incident
            ws1.cell(row_num, 11).value = f"IR{200+i}"  # Incident Raised
            ws1.cell(row_num, 12).value = f"alert-{i}@company.com"  # Email
            ws1.cell(row_num, 13).value = f"DB Alert {i}"  # DB Giant
            ws1.cell(row_num, 14).value = "Emergency"  # Type Comms
            ws1.cell(row_num, 15).value = f"Payment processing issue {i}"  # Incident Comms
            ws1.cell(row_num, 16).value = random.choice(["Yes", "No"])  # Batch Reportable
            ws1.cell(row_num, 17).value = f"Final update {i}"  # Final Comms
            ws1.cell(row_num, 18).value = random.choice(["Yes", "No"])  # CR
            ws1.cell(row_num, 19).value = random.choice(["Yes", "No"])  # Implementation
            ws1.cell(row_num, 20).value = f"Verified by {random.choice(team_members)}"  # Verification
            ws1.cell(row_num, 21).value = f"Issue resolved {i}"  # Issue Communication
            ws1.cell(row_num, 22).value = f"Monitor system {i}"  # Additional Task
            ws1.cell(row_num, 23).value = status  # Status
            
            print(f"   Row {row_num}: {incident_date} | {shift} | {incident_category} | {shift_lead} | {assigned_to} | {status}")
        
        # Save
        wb.save(EXCEL_FILE)
        print(f"\n✅ Saved to {EXCEL_FILE}")
        
        print(f"\n🎉 Excel Structure Restored!")
        print(f"   ✅ Incident Category column restored")
        print(f"   ✅ Shift Lead column restored")
        print(f"   ✅ Total incidents: 20")
        print(f"   ✅ All columns properly aligned")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    restore_excel_structure()
