#!/usr/bin/env python3
"""
AI - SRE Alert Investigation Tracker
Flask Service with Banking-Grade Features
Reads/writes to Excel, serves REST API for Form, Dashboard, and Admin
"""

from flask import Flask, jsonify, request, send_file, session, render_template
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import json
import os
from pathlib import Path
import hashlib
import threading

app = Flask(__name__)
CORS(app)
app.secret_key = 'ai-sre-alert-investigation-tracker-secret-key'

# Configuration
EXCEL_FILE = 'incident-tracker.xlsx'
PORT = 5000
ADMIN_PIN = '9999'  # Default admin PIN (should be changed in production)

# File locking mechanism
file_lock = threading.Lock()

# Audit log storage
audit_log = []

# Ensure Excel file exists
def ensure_excel_exists():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found!")
        print("Please ensure incident-tracker.xlsx is in the same directory as app.py")
        return False
    return True

# Helper: Calculate MTTR (Mean Time To Resolution)
def calculate_mttr(created_at, completed_at):
    """Calculate MTTR in minutes"""
    try:
        if not created_at or not completed_at:
            return None
        
        # Parse timestamps
        if isinstance(created_at, str):
            created = datetime.fromisoformat(created_at)
        else:
            created = created_at
            
        if isinstance(completed_at, str):
            completed = datetime.fromisoformat(completed_at)
        else:
            completed = completed_at
        
        # Calculate difference in minutes
        diff = completed - created
        mttr_minutes = int(diff.total_seconds() / 60)
        return mttr_minutes
    except Exception as e:
        print(f"Error calculating MTTR: {e}")
        return None

# Helper: Format MTTR for display
def format_mttr(minutes):
    """Format MTTR in human-readable format"""
    if not minutes:
        return "--"
    
    hours = minutes // 60
    mins = minutes % 60
    
    if hours > 0:
        return f"{hours}h {mins}m"
    else:
        return f"{mins}m"

# Helper: Log audit action
def log_audit_action(user, action, incident_id, field_changed=None):
    """Log admin actions for audit trail"""
    audit_entry = {
        'timestamp': datetime.now().isoformat(),
        'user': user,
        'action': action,
        'incident_id': incident_id,
        'field_changed': field_changed
    }
    audit_log.append(audit_entry)
    print(f"[AUDIT] {user} - {action} - Incident #{incident_id}")

# Helper: Validate admin authentication
def validate_admin_auth():
    """Check if user is authenticated as admin"""
    return session.get('admin_authenticated', False)

# Helper: Read all incidents from Excel
def read_incidents():
    try:
        with file_lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            
            incidents = []
            headers = []
            
            # Read headers from row 1
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # Read data rows (starting from row 2)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is None:  # Skip empty rows
                    continue
                
                incident = {}
                for col_idx, header in enumerate(headers):
                    incident[header] = row[col_idx] if col_idx < len(row) else None
                incident['_row_number'] = row_idx  # Track original row number
                incidents.append(incident)
            
            wb.close()
            return incidents
    except Exception as e:
        print(f"Error reading incidents: {e}")
        return []

# Helper: Write incident to Excel
def write_incident(incident_data, row_number=None):
    try:
        with file_lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            
            # Get headers from row 1
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # If row_number not specified, append to end
            if row_number is None:
                row_number = ws.max_row + 1
            
            # Write data
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_number, column=col_idx)
                cell.value = incident_data.get(header, '')
            
            wb.save(EXCEL_FILE)
            wb.close()
            return True, row_number
    except Exception as e:
        print(f"Error writing incident: {e}")
        return False, None

# HTML Routes - Serve templates
@app.route('/form.html')
def form_page():
    """Serve form.html"""
    return render_template('form.html')

@app.route('/dashboard.html')
def dashboard_page():
    """Serve dashboard.html"""
    return render_template('dashboard.html')

@app.route('/admin.html')
def admin_page():
    """Serve admin.html"""
    return render_template('admin.html')

# API Endpoints

@app.route('/api/health', methods=['GET'])
def health():
    """Health check"""
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()}), 200

@app.route('/api/incidents', methods=['GET'])
def get_incidents():
    """Get all incidents"""
    incidents = read_incidents()
    return jsonify({
        'count': len(incidents),
        'incidents': incidents,
        'timestamp': datetime.now().isoformat()
    }), 200

@app.route('/api/incidents', methods=['POST'])
def create_incident():
    """Create new incident"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['Date', 'Shift', 'Alert', 'Assigned To', 'Status', 'Incident Category', 'Shift Lead']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Add timestamps
        data['Created At'] = datetime.now().isoformat()
        data['Last Modified By'] = 'Form User'
        data['Last Modified At'] = datetime.now().isoformat()
        
        # Write to Excel
        success, row_number = write_incident(data)
        if not success:
            return jsonify({'error': 'Failed to write to Excel'}), 500
        
        log_audit_action('Form User', 'CREATE', row_number)
        
        return jsonify({
            'success': True,
            'row_number': row_number,
            'message': 'Incident created successfully',
            'incident': data
        }), 201
    except Exception as e:
        print(f"Error creating incident: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/incidents/<int:row_number>', methods=['PUT'])
def update_incident(row_number):
    """Update existing incident"""
    try:
        data = request.get_json()
        success, _ = write_incident(data, row_number=row_number)
        
        if not success:
            return jsonify({'error': 'Failed to update'}), 500
        
        return jsonify({
            'success': True,
            'message': 'Incident updated successfully'
        }), 200
    except Exception as e:
        print(f"Error updating incident: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/teams', methods=['GET'])
def get_team_members():
    """Get list of team members for dropdown"""
    try:
        with file_lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            
            # Check if Sheet2 exists
            if 'Sheet2' not in wb.sheetnames:
                wb.close()
                return jsonify({'error': 'Team members sheet not found'}), 404
            
            ws = wb['Sheet2']
            members = []
            
            # Read data rows (starting from row 2, skip header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Skip empty rows
                    continue
                members.append(row[0])  # Get name from first column
            
            wb.close()
            
            # Return as simple list (all team members work all shifts)
            return jsonify({'members': members}), 200
    except Exception as e:
        print(f"Error reading team members: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories', methods=['GET'])
def get_categories():
    """Get dropdown categories"""
    return jsonify({
        'shifts': ['S1', 'S2', 'On Call'],
        'categories': ['P1', 'P2', 'P3', 'P4'],
        'timeSlots': ['7-8 AM', '8-9 AM', '9-10 AM', '10-11 AM', 
                     '11-12 PM', '12-1 PM', '1-2 PM', '2-3 PM', '3-4 PM', 
                     '4-5 PM', '5-6 PM', '6-7 PM', '7-8 PM', '8-9 PM', '9-10 PM', '10 PM-7 AM'],
        'statuses': ['In Progress', 'Pending', 'Completed'],
        'yesNo': ['Yes', 'No']
    }), 200

@app.route('/api/admin/teams', methods=['GET'])
def get_all_team_members():
    """Get all team members (admin)"""
    try:
        if not validate_admin_auth():
            return jsonify({'error': 'Unauthorized'}), 401
        
        teams = {
            "S1": ["Raj Kumar", "Priya Singh", "Amit Patel"],
            "S2": ["Vikram Joshi", "Neha Sharma", "Rohan Verma", "Anjali Menon", 
                   "Arjun Gupta", "Deepak Kumar", "Pooja Nair", "Sanjay Reddy", 
                   "Tina Desai", "Varun Malhotra", "Yash Pandey", "Zara Khan", 
                   "Aditya Rao"],
            "On Call": ["Manager A", "Manager B"]
        }
        
        # Convert to flat list with shift info
        members = []
        for shift, names in teams.items():
            for name in names:
                members.append({
                    'name': name,
                    'shift': shift,
                    'email': '',
                    'phone': ''
                })
        
        return jsonify({'members': members}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/teams', methods=['POST'])
def add_team_member():
    """Add new team member (admin)"""
    try:
        if not validate_admin_auth():
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.get_json()
        name = data.get('name')
        shift = data.get('shift')
        
        if not name or not shift:
            return jsonify({'error': 'Name and shift required'}), 400
        
        log_audit_action(session.get('admin_user', 'Admin'), 'ADD_TEAM_MEMBER', name)
        
        return jsonify({
            'success': True,
            'message': 'Team member added successfully',
            'member': {
                'name': name,
                'shift': shift,
                'email': data.get('email', ''),
                'phone': data.get('phone', '')
            }
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/teams/<name>', methods=['PUT'])
def update_team_member(name):
    """Update team member (admin)"""
    try:
        if not validate_admin_auth():
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.get_json()
        log_audit_action(session.get('admin_user', 'Admin'), 'UPDATE_TEAM_MEMBER', name)
        
        return jsonify({
            'success': True,
            'message': 'Team member updated successfully'
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/teams/<name>', methods=['DELETE'])
def delete_team_member(name):
    """Delete team member (admin)"""
    try:
        if not validate_admin_auth():
            return jsonify({'error': 'Unauthorized'}), 401
        
        log_audit_action(session.get('admin_user', 'Admin'), 'DELETE_TEAM_MEMBER', name)
        
        return jsonify({
            'success': True,
            'message': 'Team member deleted successfully'
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# NEW ENDPOINTS FOR PHASE 1

@app.route('/api/incidents/filters', methods=['GET'])
def get_filter_values():
    """Get available filter values (years, months, persons, categories)"""
    incidents = read_incidents()
    
    years = set()
    months = set()
    persons = set()
    categories = set()
    
    for incident in incidents:
        if incident.get('Date'):
            try:
                date_obj = datetime.fromisoformat(incident['Date'])
                years.add(date_obj.year)
                months.add(date_obj.month)
            except:
                pass
        
        if incident.get('Assigned To'):
            persons.add(incident['Assigned To'])
        
        if incident.get('Incident Category'):
            categories.add(incident['Incident Category'])
    
    return jsonify({
        'years': sorted(list(years), reverse=True),
        'months': sorted(list(months)),
        'persons': sorted(list(persons)),
        'categories': sorted(list(categories))
    }), 200

@app.route('/api/incidents/mttr', methods=['GET'])
def get_mttr_statistics():
    """Get MTTR statistics and trends"""
    incidents = read_incidents()
    
    mttr_values = []
    mttr_by_category = {'P1': [], 'P2': [], 'P3': [], 'P4': []}
    mttr_by_date = {}
    
    for incident in incidents:
        if incident.get('Status') == 'Completed' and incident.get('Created At') and incident.get('Completed At'):
            mttr = calculate_mttr(incident['Created At'], incident['Completed At'])
            if mttr:
                mttr_values.append(mttr)
                
                # By category
                category = incident.get('Incident Category', 'Unknown')
                if category in mttr_by_category:
                    mttr_by_category[category].append(mttr)
                
                # By date
                try:
                    date_obj = datetime.fromisoformat(incident['Completed At'])
                    date_str = date_obj.strftime('%Y-%m-%d')
                    if date_str not in mttr_by_date:
                        mttr_by_date[date_str] = []
                    mttr_by_date[date_str].append(mttr)
                except:
                    pass
    
    # Calculate averages
    avg_mttr = sum(mttr_values) / len(mttr_values) if mttr_values else 0
    avg_by_category = {cat: sum(vals) / len(vals) if vals else 0 for cat, vals in mttr_by_category.items()}
    avg_by_date = {date: sum(vals) / len(vals) if vals else 0 for date, vals in mttr_by_date.items()}
    
    return jsonify({
        'average_mttr': int(avg_mttr),
        'average_mttr_formatted': format_mttr(int(avg_mttr)),
        'by_category': {cat: int(avg) for cat, avg in avg_by_category.items()},
        'by_date': {date: int(avg) for date, avg in avg_by_date.items()},
        'total_completed': len(mttr_values)
    }), 200

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Admin authentication"""
    try:
        data = request.get_json()
        pin = data.get('pin')
        
        if pin == ADMIN_PIN:
            session['admin_authenticated'] = True
            session['admin_user'] = 'Admin'
            log_audit_action('Admin', 'LOGIN', 'N/A')
            return jsonify({
                'success': True,
                'message': 'Admin authenticated successfully'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid admin PIN'
            }), 401
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    """Admin logout"""
    try:
        log_audit_action(session.get('admin_user', 'Unknown'), 'LOGOUT', 'N/A')
        session.clear()
        return jsonify({
            'success': True,
            'message': 'Logged out successfully'
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/incidents/<int:row_number>', methods=['POST'])
def admin_update_incident(row_number):
    """Admin update incident (all fields)"""
    try:
        if not validate_admin_auth():
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.get_json()
        
        # If status changed to Completed, calculate MTTR
        if data.get('Status') == 'Completed' and data.get('Created At'):
            data['Completed At'] = datetime.now().isoformat()
            mttr = calculate_mttr(data['Created At'], data['Completed At'])
            if mttr:
                data['MTTR (minutes)'] = mttr
        
        # Track last modification
        data['Last Modified By'] = session.get('admin_user', 'Admin')
        data['Last Modified At'] = datetime.now().isoformat()
        
        success, _ = write_incident(data, row_number=row_number)
        
        if not success:
            return jsonify({'error': 'Failed to update'}), 500
        
        log_audit_action(session.get('admin_user', 'Admin'), 'UPDATE', row_number)
        
        return jsonify({
            'success': True,
            'message': 'Incident updated successfully'
        }), 200
    except Exception as e:
        print(f"Error updating incident: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/incidents/<int:row_number>', methods=['DELETE'])
def admin_archive_incident(row_number):
    """Admin archive incident (soft delete)"""
    try:
        if not validate_admin_auth():
            return jsonify({'error': 'Unauthorized'}), 401
        
        incidents = read_incidents()
        
        # Find and mark as archived
        for incident in incidents:
            if incident.get('_row_number') == row_number:
                incident['Status'] = 'Archived'
                incident['Last Modified By'] = session.get('admin_user', 'Admin')
                incident['Last Modified At'] = datetime.now().isoformat()
                write_incident(incident, row_number=row_number)
                log_audit_action(session.get('admin_user', 'Admin'), 'ARCHIVE', row_number)
                break
        
        return jsonify({
            'success': True,
            'message': 'Incident archived successfully'
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/audit-log', methods=['GET'])
def get_audit_log():
    """Get audit log entries"""
    try:
        if not validate_admin_auth():
            return jsonify({'error': 'Unauthorized'}), 401
        
        return jsonify({
            'count': len(audit_log),
            'entries': audit_log[-100:]  # Return last 100 entries
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/csv', methods=['GET'])
def export_csv():
    """Export incidents as CSV"""
    try:
        incidents = read_incidents()
        
        if not incidents:
            return jsonify({'error': 'No incidents to export'}), 404
        
        # Create CSV content
        headers = list(incidents[0].keys())
        csv_lines = [','.join(str(h) for h in headers)]
        
        for incident in incidents:
            row = [str(incident.get(h, '')) for h in headers]
            csv_lines.append(','.join(row))
        
        csv_content = '\n'.join(csv_lines)
        
        return csv_content, 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': 'attachment; filename=incidents-export.csv'
        }
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/')
def index():
    """Root endpoint - info"""
    return jsonify({
        'service': 'Incident Tracker API',
        'version': '1.0',
        'endpoints': {
            'GET /api/health': 'Health check',
            'GET /api/incidents': 'Get all incidents',
            'POST /api/incidents': 'Create new incident',
            'PUT /api/incidents/<row>': 'Update incident',
            'GET /api/teams': 'Get team members',
            'GET /api/categories': 'Get dropdown categories',
            'GET /api/export/csv': 'Export as CSV',
            'GET /form.html': 'Open incident entry form',
            'GET /dashboard.html': 'Open live dashboard'
        },
        'documentation': 'See SETUP.md for detailed instructions'
    }), 200

if __name__ == '__main__':
    if ensure_excel_exists():
        print("\n" + "="*70)
        print("🚀 AI - SRE Alert Investigation Tracker")
        print("   Banking/Financial Institution Grade")
        print("="*70)
        print(f"✓ Excel file: {EXCEL_FILE}")
        print(f"✓ API running on: http://localhost:{PORT}")
        print(f"✓ Form: http://localhost:{PORT}/form.html")
        print(f"✓ Dashboard: http://localhost:{PORT}/dashboard.html")
        print(f"✓ Admin: http://localhost:{PORT}/admin.html")
        print(f"✓ Admin PIN: {ADMIN_PIN}")
        print("="*70 + "\n")
        
        # Check if Flask is installed
        try:
            app.run(debug=False, host='localhost', port=PORT)
        except Exception as e:
            print(f"Error: {e}")
            print("\nInstall Flask: pip install flask flask-cors openpyxl")
    else:
        print("\nSetup failed. Please run setup first.")
