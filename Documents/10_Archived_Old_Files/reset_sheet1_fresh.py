#!/usr/bin/env python3
"""
Reset Sheet1 - Delete all dummy data and create fresh entries with new team members
"""

import openpyxl
from datetime import datetime, timedelta
import random

EXCEL_FILE = 'incident-tracker.xlsx'

def get_new_team_members():
    """Read new team members from Sheet2"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Sheet2']
        members = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                members.append(row[0])
        
        wb.close()
        return members
    except Exception as e:
        print(f"❌ Error reading Sheet2: {e}")
        return []

def reset_sheet1():
    """Reset Sheet1 with fresh dummy data"""
    try:
        # Get new team members
        new_team = get_new_team_members()
        
        if not new_team:
            print("❌ No team members found in Sheet2!")
            return
        
        print(f"📝 Found {len(new_team)} team members from Sheet2")
        print(f"   {', '.join(new_team[:5])}...\n")
        
        # Open workbook
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Sheet1']
        
        # Keep only header row, delete all data rows
        print("🗑️  Deleting all existing dummy data...")
        while ws.max_row > 1:
            ws.delete_rows(2, 1)
        
        print("✅ All dummy data deleted\n")
        
        # Create fresh dummy data with new team members
        print("📊 Creating fresh dummy data with new team members...\n")
        
        base_date = datetime.now()
        shifts = ["S1", "S2", "On Call"]
        statuses = ["In Progress", "Pending", "Completed"]
        categories = ["P1", "P2", "P3", "P4"]
        time_slots = ["7-8 AM", "8-9 AM", "9-10 AM", "10-11 AM", "11-12 PM", "12-1 PM", 
                     "1-2 PM", "2-3 PM", "3-4 PM", "4-5 PM", "5-6 PM", "6-7 PM", 
                     "7-8 PM", "8-9 PM", "9-10 PM", "10 PM-7 AM"]
        
        # Create 20 fresh incidents
        for i in range(20):
            row_num = i + 2
            incident_date = (base_date - timedelta(days=i)).strftime('%Y-%m-%d')
            shift = random.choice(shifts)
            shift_lead = random.choice(new_team)
            assigned_to = random.choice(new_team)
            status = random.choice(statuses)
            category = random.choice(categories)
            time_slot = random.choice(time_slots)
            alert_time = f"{random.randint(6, 22):02d}:{random.randint(0, 59):02d}"
            
            # Write to row
            ws.cell(row_num, 1).value = incident_date  # Date
            ws.cell(row_num, 2).value = shift  # Shift
            ws.cell(row_num, 3).value = shift_lead  # Shift Lead
            ws.cell(row_num, 4).value = time_slot  # Time Slot
            ws.cell(row_num, 5).value = alert_time  # Alert Report Time
            ws.cell(row_num, 6).value = f"Payment Gateway Timeout - Transaction {1000+i}"  # Alert
            ws.cell(row_num, 7).value = assigned_to  # Assigned To
            ws.cell(row_num, 8).value = f"INC{1000+i}"  # RITM
            ws.cell(row_num, 9).value = f"STIP{500+i}"  # STIP Incident
            ws.cell(row_num, 10).value = f"IR{200+i}"  # Incident Raised
            ws.cell(row_num, 11).value = f"alert-{i}@company.com"  # Email
            ws.cell(row_num, 12).value = f"DB Alert {i}"  # DB Giant
            ws.cell(row_num, 13).value = "Emergency"  # Type Comms
            ws.cell(row_num, 14).value = f"Payment processing issue {i}"  # Incident Comms
            ws.cell(row_num, 15).value = random.choice(["Yes", "No"])  # Batch Reportable
            ws.cell(row_num, 16).value = f"Final update {i}"  # Final Comms
            ws.cell(row_num, 17).value = random.choice(["Yes", "No"])  # CR
            ws.cell(row_num, 18).value = random.choice(["Yes", "No"])  # Implementation
            ws.cell(row_num, 19).value = f"Verified by {random.choice(new_team)}"  # Verification
            ws.cell(row_num, 20).value = f"Issue resolved {i}"  # Issue Communication
            ws.cell(row_num, 21).value = f"Monitor system {i}"  # Additional Task
            ws.cell(row_num, 22).value = status  # Status
            ws.cell(row_num, 23).value = category  # Incident Category
            
            print(f"   Row {row_num}: {incident_date} | {shift} | {shift_lead} → {assigned_to} | {status}")
        
        # Save
        wb.save(EXCEL_FILE)
        print(f"\n✅ Saved to {EXCEL_FILE}")
        
        print(f"\n🎉 Sheet1 Reset Complete!")
        print(f"   Total new incidents: 20")
        print(f"   Team members used: {len(new_team)}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    reset_sheet1()
