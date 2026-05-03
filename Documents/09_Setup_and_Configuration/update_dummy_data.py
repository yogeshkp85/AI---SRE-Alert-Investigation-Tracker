#!/usr/bin/env python3
"""
Update dummy incident data with new team member names
"""

import openpyxl
from datetime import datetime, timedelta
import random

EXCEL_FILE = 'incident-tracker.xlsx'

# New team member list
TEAM_MEMBERS = [
    "Raj Kumar",
    "Priya Singh",
    "Amit Patel",
    "Vikram Joshi",
    "Neha Sharma",
    "Rohan Verma",
    "Anjali Menon",
    "Arjun Gupta",
    "Deepak Kumar",
    "Pooja Nair",
    "Sanjay Reddy",
    "Tina Desai",
    "Varun Malhotra",
    "Yash Pandey",
    "Zara Khan",
    "Aditya Rao",
    "Manager A",
    "Manager B"
]

SHIFTS = ["S1", "S2", "On Call"]
STATUSES = ["In Progress", "Pending", "Completed"]
CATEGORIES = ["P1", "P2", "P3", "P4"]

def update_dummy_data():
    """Update existing incidents with new team member names"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1] if cell.value]
        
        # Find column indices
        shift_lead_col = headers.index("Shift Lead") + 1 if "Shift Lead" in headers else None
        assigned_to_col = headers.index("Assigned To") + 1 if "Assigned To" in headers else None
        
        print(f"📝 Updating dummy data...")
        print(f"   Shift Lead Column: {shift_lead_col}")
        print(f"   Assigned To Column: {assigned_to_col}")
        
        # Update existing rows (starting from row 2)
        updated_count = 0
        for row_idx in range(2, ws.max_row + 1):
            # Check if row has data
            if ws.cell(row_idx, 1).value is None:
                continue
            
            # Update Shift Lead
            if shift_lead_col:
                ws.cell(row_idx, shift_lead_col).value = random.choice(TEAM_MEMBERS)
            
            # Update Assigned To
            if assigned_to_col:
                ws.cell(row_idx, assigned_to_col).value = random.choice(TEAM_MEMBERS)
            
            updated_count += 1
        
        # Add more sample incidents with new team members
        print(f"\n✅ Updated {updated_count} existing incidents")
        print(f"📊 Adding 10 new sample incidents with new team members...")
        
        base_date = datetime.now()
        for i in range(10):
            row_num = ws.max_row + 1
            incident_date = (base_date - timedelta(days=i)).strftime('%Y-%m-%d')
            
            incident_data = {
                "Date": incident_date,
                "Shift": random.choice(SHIFTS),
                "Shift Lead": random.choice(TEAM_MEMBERS),
                "Time Slot": random.choice(["7-8 AM", "8-9 AM", "9-10 AM", "10-11 AM", "11-12 PM", "12-1 PM", "1-2 PM", "2-3 PM", "3-4 PM", "4-5 PM", "5-6 PM", "6-7 PM", "7-8 PM", "8-9 PM", "9-10 PM", "10 PM-7 AM"]),
                "Alert Report Time": f"{random.randint(6, 22):02d}:{random.randint(0, 59):02d}",
                "Alert": f"Payment Gateway Timeout - Transaction {1000+i}",
                "Assigned To": random.choice(TEAM_MEMBERS),
                "RITM": f"INC{1000+i}",
                "STIP Incident": f"STIP{500+i}",
                "Incident Raised": f"IR{200+i}",
                "Email": f"alert-{i}@company.com",
                "DB Giant": f"DB Alert {i}",
                "Type Comms": "Emergency",
                "Incident Comms": f"Payment processing issue {i}",
                "Batch Reportable": random.choice(["Yes", "No"]),
                "Final Comms": f"Final update {i}",
                "CR": random.choice(["Yes", "No"]),
                "Implementation": random.choice(["Yes", "No"]),
                "Verification": f"Verified by {random.choice(TEAM_MEMBERS)}",
                "Issue Communication": f"Issue resolved {i}",
                "Additional Task/Improvement": f"Monitor system {i}",
                "Status": random.choice(STATUSES)
            }
            
            # Write to Excel
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row_num, col_idx).value = incident_data.get(header, '')
        
        wb.save(EXCEL_FILE)
        print(f"✅ Added 10 new incidents")
        print(f"\n🎉 Dummy data updated successfully!")
        print(f"   Total incidents: {ws.max_row - 1}")
        print(f"\n📊 Team members used:")
        for member in TEAM_MEMBERS:
            print(f"   - {member}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    update_dummy_data()
