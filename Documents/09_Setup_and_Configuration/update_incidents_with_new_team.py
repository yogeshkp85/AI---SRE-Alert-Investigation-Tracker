#!/usr/bin/env python3
"""
Update incident data with new team member names from Sheet2
"""

import openpyxl
from datetime import datetime, timedelta
import random

EXCEL_FILE = 'incident-tracker.xlsx'

def get_team_members_from_sheet2():
    """Read team members from Sheet2"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        if 'Sheet2' not in wb.sheetnames:
            print("❌ Sheet2 not found!")
            return []
        
        ws = wb['Sheet2']
        members = []
        
        # Read from row 2 onwards (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If Name column has value
                members.append(row[0])
        
        wb.close()
        return members
    except Exception as e:
        print(f"❌ Error reading Sheet2: {e}")
        return []

def update_incidents_with_new_team():
    """Update incidents with new team member names"""
    try:
        # Get new team members
        new_team = get_team_members_from_sheet2()
        
        if not new_team:
            print("❌ No team members found in Sheet2!")
            return
        
        print(f"📝 Found {len(new_team)} team members in Sheet2:")
        for member in new_team:
            print(f"   - {member}")
        
        # Open workbook
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # Get the Incidents sheet (first sheet)
        if 'Incidents' in wb.sheetnames:
            ws = wb['Incidents']
        else:
            ws = wb.worksheets[0]  # First sheet
        
        # Get headers
        headers = [cell.value for cell in ws[1] if cell.value]
        print(f"\n📊 Headers: {headers}")
        
        # Find column indices
        shift_lead_col = None
        assigned_to_col = None
        
        for idx, header in enumerate(headers, 1):
            if header == "Shift Lead":
                shift_lead_col = idx
            elif header == "Assigned To":
                assigned_to_col = idx
        
        print(f"   Shift Lead Column: {shift_lead_col}")
        print(f"   Assigned To Column: {assigned_to_col}")
        
        # Update existing incidents
        updated_count = 0
        for row_idx in range(2, ws.max_row + 1):
            # Check if row has data
            if ws.cell(row_idx, 1).value is None:
                continue
            
            # Update Shift Lead
            if shift_lead_col:
                ws.cell(row_idx, shift_lead_col).value = random.choice(new_team)
            
            # Update Assigned To
            if assigned_to_col:
                ws.cell(row_idx, assigned_to_col).value = random.choice(new_team)
            
            updated_count += 1
        
        print(f"\n✅ Updated {updated_count} existing incidents with new team members")
        
        # Save
        wb.save(EXCEL_FILE)
        print(f"✅ Saved to {EXCEL_FILE}")
        
        print(f"\n🎉 All incidents updated successfully!")
        print(f"   Total incidents: {updated_count}")
        print(f"   Team members used: {len(new_team)}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    update_incidents_with_new_team()
