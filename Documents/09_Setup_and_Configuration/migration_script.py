#!/usr/bin/env python3
"""
Migration Script - Add 7 new columns to incident-tracker.xlsx
Columns: Incident Category, Shift Lead, Created At, Completed At, MTTR (minutes), Last Modified By, Last Modified At
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import shutil
import os

EXCEL_FILE = 'incident-tracker.xlsx'
BACKUP_FILE = 'incident-tracker.backup.xlsx'

def migrate_excel():
    """Add new columns to Excel file"""
    
    # Create backup
    if os.path.exists(EXCEL_FILE):
        shutil.copy(EXCEL_FILE, BACKUP_FILE)
        print(f"✓ Backup created: {BACKUP_FILE}")
    else:
        print(f"Error: {EXCEL_FILE} not found!")
        return False
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Get current headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        print(f"Current columns: {len(headers)}")
        print(f"Headers: {headers}")
        
        # New columns to add
        new_columns = [
            'Incident Category',
            'Shift Lead',
            'Created At',
            'Completed At',
            'MTTR (minutes)',
            'Last Modified By',
            'Last Modified At'
        ]
        
        # Add new headers
        next_col = len(headers) + 1
        for idx, col_name in enumerate(new_columns, start=next_col):
            cell = ws.cell(row=1, column=idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Update existing rows with default values
        for row_idx in range(2, ws.max_row + 1):
            # Incident Category (default: P4)
            ws.cell(row=row_idx, column=next_col).value = 'P4'
            
            # Shift Lead (default: TBD)
            ws.cell(row=row_idx, column=next_col + 1).value = 'TBD'
            
            # Created At (default: current date)
            ws.cell(row=row_idx, column=next_col + 2).value = datetime.now().isoformat()
            
            # Completed At (default: empty)
            ws.cell(row=row_idx, column=next_col + 3).value = ''
            
            # MTTR (minutes) (default: 0)
            ws.cell(row=row_idx, column=next_col + 4).value = 0
            
            # Last Modified By (default: System)
            ws.cell(row=row_idx, column=next_col + 5).value = 'System'
            
            # Last Modified At (default: current date)
            ws.cell(row=row_idx, column=next_col + 6).value = datetime.now().isoformat()
        
        # Save workbook
        wb.save(EXCEL_FILE)
        wb.close()
        
        print(f"✓ Excel file updated: {EXCEL_FILE}")
        print(f"✓ Added {len(new_columns)} new columns")
        print(f"✓ Updated {ws.max_row - 1} existing rows")
        return True
        
    except Exception as e:
        print(f"❌ Migration failed: {e}")
        return False

if __name__ == '__main__':
    print("\n" + "="*70)
    print("🚀 Excel Migration Script")
    print("="*70 + "\n")
    
    if migrate_excel():
        print("\n✓ Migration completed successfully!")
        print("="*70 + "\n")
    else:
        print("\n❌ Migration failed!")
        print("="*70 + "\n")
