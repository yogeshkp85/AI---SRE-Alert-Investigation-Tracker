#!/usr/bin/env python3
"""
Populate dummy data into incident-tracker.xlsx for testing
Creates 25 sample incidents with realistic data
"""

import openpyxl
from datetime import datetime, timedelta
import random

EXCEL_FILE = 'incident-tracker.xlsx'

# Sample data
SHIFTS = ['S1', 'S2', 'On Call']
TIME_SLOTS = {
    'S1': ['7-8 AM', '8-9 AM', '9-10 AM', '10-11 AM'],
    'S2': ['11-12 PM', '12-1 PM', '1-2 PM', '2-3 PM', '3-4 PM', '4-5 PM', '5-6 PM', '6-7 PM', '7-8 PM', '8-9 PM', '9-10 PM'],
    'On Call': ['10 PM-7 AM']
}
TIMES = ['08:30', '09:15', '10:45', '11:20', '13:00', '14:30', '15:45', '16:20', '17:00', '18:30']
ALERTS = [
    'Database connection timeout',
    'Payment gateway API failure',
    'High memory usage detected',
    'Network latency spike',
    'Service unavailable error',
    'Transaction processing delay',
    'Authentication service down',
    'Cache invalidation issue',
    'Load balancer error',
    'SSL certificate expiration warning'
]
TEAM_MEMBERS = ['Raj Kumar', 'Priya Singh', 'Amit Patel', 'Vikram Joshi', 'Neha Sharma', 
                'Rohan Verma', 'Anjali Menon', 'Arjun Gupta', 'Deepak Kumar', 'Pooja Nair']
CATEGORIES = ['P1', 'P2', 'P3', 'P4']
STATUSES = ['In Progress', 'Pending', 'Completed']
RITM_REFS = ['RIT0001234', 'RIT0001235', 'RIT0001236', 'RIT0001237', 'RIT0001238']
VERIFICATION_NOTES = [
    'Issue resolved after cache clear',
    'Temporary fix applied, permanent solution pending',
    'Root cause identified and documented',
    'Monitoring in place for recurrence',
    'Escalated to infrastructure team'
]

def generate_dummy_data():
    """Generate 25 dummy incident entries"""
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        print(f"Loading {EXCEL_FILE}...")
        print(f"Current rows: {ws.max_row}")
        
        # Start from row 2 (row 1 is headers)
        start_row = 2
        
        # Generate 25 dummy entries
        for i in range(25):
            row = start_row + i
            
            # Select random values
            shift = random.choice(SHIFTS)
            time_slot = random.choice(TIME_SLOTS[shift])
            date = (datetime.now() - timedelta(days=random.randint(0, 30))).strftime('%Y-%m-%d')
            time = random.choice(TIMES)
            alert = random.choice(ALERTS)
            assigned_to = random.choice(TEAM_MEMBERS)
            shift_lead = random.choice(TEAM_MEMBERS)
            category = random.choice(CATEGORIES)
            status = random.choice(STATUSES)
            ritm = random.choice(RITM_REFS)
            verification = random.choice(VERIFICATION_NOTES)
            
            # Write data to columns
            ws.cell(row=row, column=1).value = date                    # Date
            ws.cell(row=row, column=2).value = shift                   # Shift
            ws.cell(row=row, column=3).value = time_slot               # Time Slot
            ws.cell(row=row, column=4).value = time                    # Alert Report Time
            ws.cell(row=row, column=5).value = alert                   # Alert
            ws.cell(row=row, column=6).value = assigned_to             # Assigned To
            ws.cell(row=row, column=7).value = ritm                    # RITM
            ws.cell(row=row, column=8).value = f'STIP{i:05d}'          # STIP Incident
            ws.cell(row=row, column=9).value = f'INC{i:05d}'           # Incident Raised
            ws.cell(row=row, column=10).value = f'Email-{i}'           # Email
            ws.cell(row=row, column=11).value = f'DB-{i}'              # DB Giant
            ws.cell(row=row, column=12).value = 'Email'                # Type Comms
            ws.cell(row=row, column=13).value = 'Incident Alert'       # Incident Comms
            ws.cell(row=row, column=14).value = random.choice(['Yes', 'No'])  # Batch Reportable
            ws.cell(row=row, column=15).value = f'Final-{i}'           # Final Comms
            ws.cell(row=row, column=16).value = random.choice(['Yes', 'No'])  # CR
            ws.cell(row=row, column=17).value = random.choice(['Yes', 'No'])  # Implementation
            ws.cell(row=row, column=18).value = verification           # Verification
            ws.cell(row=row, column=19).value = 'Issue resolved'       # Issue Communication
            ws.cell(row=row, column=20).value = 'Monitor for recurrence'  # Additional Task/Improvement
            ws.cell(row=row, column=21).value = status                 # Status
            
            # New columns (if they exist)
            if ws.max_column >= 22:
                ws.cell(row=row, column=22).value = category           # Incident Category
            if ws.max_column >= 23:
                ws.cell(row=row, column=23).value = shift_lead         # Shift Lead
            if ws.max_column >= 24:
                ws.cell(row=row, column=24).value = datetime.now().isoformat()  # Created At
            if ws.max_column >= 25:
                ws.cell(row=row, column=25).value = ''                 # Completed At
            if ws.max_column >= 26:
                ws.cell(row=row, column=26).value = 0                  # MTTR (minutes)
            if ws.max_column >= 27:
                ws.cell(row=row, column=27).value = 'Test User'        # Last Modified By
            if ws.max_column >= 28:
                ws.cell(row=row, column=28).value = datetime.now().isoformat()  # Last Modified At
            
            print(f"✓ Row {row}: {date} | {shift} | {alert[:30]}... | {status}")
        
        # Save workbook
        wb.save(EXCEL_FILE)
        wb.close()
        
        print(f"\n✓ Successfully added 25 dummy entries!")
        print(f"✓ File saved: {EXCEL_FILE}")
        print(f"\nYou can now:")
        print("1. Start the Flask app: python app.py")
        print("2. Open dashboard: http://localhost:5000/dashboard.html")
        print("3. View the 25 test incidents")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False
    
    return True

if __name__ == '__main__':
    print("\n" + "="*70)
    print("🚀 Populate Dummy Data Script")
    print("="*70 + "\n")
    
    generate_dummy_data()
    
    print("\n" + "="*70 + "\n")
