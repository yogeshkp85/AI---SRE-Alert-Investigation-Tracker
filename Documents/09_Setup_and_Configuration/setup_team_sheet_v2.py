#!/usr/bin/env python3
"""
Setup Team Members Sheet in Excel (Version 2)
Creates Sheet2 with team member data - NO SHIFT COLUMN (everyone works all shifts)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_FILE = 'incident-tracker.xlsx'

# Team members data - NO SHIFT COLUMN (everyone works in rotation)
TEAM_MEMBERS = [
    ('Raj Kumar', 'raj.kumar@company.com', '9876543210'),
    ('Priya Singh', 'priya.singh@company.com', '9876543211'),
    ('Amit Patel', 'amit.patel@company.com', '9876543212'),
    ('Vikram Joshi', 'vikram.joshi@company.com', '9876543213'),
    ('Neha Sharma', 'neha.sharma@company.com', '9876543214'),
    ('Rohan Verma', 'rohan.verma@company.com', '9876543215'),
    ('Anjali Menon', 'anjali.menon@company.com', '9876543216'),
    ('Arjun Gupta', 'arjun.gupta@company.com', '9876543217'),
    ('Deepak Kumar', 'deepak.kumar@company.com', '9876543218'),
    ('Pooja Nair', 'pooja.nair@company.com', '9876543219'),
    ('Sanjay Reddy', 'sanjay.reddy@company.com', '9876543220'),
    ('Tina Desai', 'tina.desai@company.com', '9876543221'),
    ('Varun Malhotra', 'varun.malhotra@company.com', '9876543222'),
    ('Yash Pandey', 'yash.pandey@company.com', '9876543223'),
    ('Zara Khan', 'zara.khan@company.com', '9876543224'),
    ('Aditya Rao', 'aditya.rao@company.com', '9876543225'),
    ('Manager A', 'manager.a@company.com', '9876543226'),
    ('Manager B', 'manager.b@company.com', '9876543227'),
]

print("\n" + "="*80)
print("🔧 SETTING UP TEAM MEMBERS SHEET (V2 - NO SHIFT COLUMN)")
print("="*80)

try:
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    print(f"\n✓ Loaded workbook: {EXCEL_FILE}")
    print(f"  Current sheets: {wb.sheetnames}")
    
    # Remove Sheet2 if it exists
    if 'Sheet2' in wb.sheetnames:
        print("\n⚠️  Sheet2 already exists - removing it...")
        del wb['Sheet2']
    
    # Create new Sheet2
    ws = wb.create_sheet('Sheet2', 1)
    print("✓ Created Sheet2")
    
    # Add headers (NO SHIFT COLUMN)
    headers = ['Name', 'Email', 'Phone']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print(f"✓ Added headers: {headers}")
    
    # Add team members
    for row, (name, email, phone) in enumerate(TEAM_MEMBERS, 2):
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=2).value = email
        ws.cell(row=row, column=3).value = phone
    
    print(f"✓ Added {len(TEAM_MEMBERS)} team members")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    
    # Save workbook
    wb.save(EXCEL_FILE)
    wb.close()
    
    print(f"\n✅ Successfully saved {EXCEL_FILE}")
    print("\n" + "="*80)
    print("📊 TEAM MEMBERS SHEET CREATED (NO SHIFT COLUMN)")
    print("="*80)
    print(f"\nSheet2 contains {len(TEAM_MEMBERS)} team members:")
    print(f"{'#':>3} | {'Name':<20} | {'Email':<30} | {'Phone':<12}")
    print("-" * 70)
    for i, (name, email, phone) in enumerate(TEAM_MEMBERS, 1):
        print(f"{i:3d} | {name:<20} | {email:<30} | {phone:<12}")
    
    print("\n✓ All team members work in rotation across all shifts")
    print("✓ Next step: Update backend to read/write team members from Sheet2")
    
except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n")
