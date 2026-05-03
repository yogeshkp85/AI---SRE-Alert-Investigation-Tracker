#!/usr/bin/env python3
"""
Setup Team Members Sheet in Excel
Creates Sheet2 with team member data and updates backend to use it
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_FILE = 'incident-tracker.xlsx'

# Team members data
TEAM_MEMBERS = [
    ('Raj Kumar', 'S1', 'raj.kumar@company.com', '9876543210'),
    ('Priya Singh', 'S1', 'priya.singh@company.com', '9876543211'),
    ('Amit Patel', 'S1', 'amit.patel@company.com', '9876543212'),
    ('Vikram Joshi', 'S2', 'vikram.joshi@company.com', '9876543213'),
    ('Neha Sharma', 'S2', 'neha.sharma@company.com', '9876543214'),
    ('Rohan Verma', 'S2', 'rohan.verma@company.com', '9876543215'),
    ('Anjali Menon', 'S2', 'anjali.menon@company.com', '9876543216'),
    ('Arjun Gupta', 'S2', 'arjun.gupta@company.com', '9876543217'),
    ('Deepak Kumar', 'S2', 'deepak.kumar@company.com', '9876543218'),
    ('Pooja Nair', 'S2', 'pooja.nair@company.com', '9876543219'),
    ('Sanjay Reddy', 'S2', 'sanjay.reddy@company.com', '9876543220'),
    ('Tina Desai', 'S2', 'tina.desai@company.com', '9876543221'),
    ('Varun Malhotra', 'S2', 'varun.malhotra@company.com', '9876543222'),
    ('Yash Pandey', 'S2', 'yash.pandey@company.com', '9876543223'),
    ('Zara Khan', 'S2', 'zara.khan@company.com', '9876543224'),
    ('Aditya Rao', 'S2', 'aditya.rao@company.com', '9876543225'),
    ('Manager A', 'On Call', 'manager.a@company.com', '9876543226'),
    ('Manager B', 'On Call', 'manager.b@company.com', '9876543227'),
]

print("\n" + "="*80)
print("🔧 SETTING UP TEAM MEMBERS SHEET")
print("="*80)

try:
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    print(f"\n✓ Loaded workbook: {EXCEL_FILE}")
    print(f"  Current sheets: {wb.sheetnames}")
    
    # Remove Sheet2 if it exists
    if 'Sheet2' in wb.sheetnames:
        print("\n⚠️  Sheet2 already exists - removing it...")
        del wb['Sheet2']
    
    # Create new Sheet2
    ws = wb.create_sheet('Sheet2', 1)
    print("✓ Created Sheet2")
    
    # Add headers
    headers = ['Name', 'Shift', 'Email', 'Phone']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print(f"✓ Added headers: {headers}")
    
    # Add team members
    for row, (name, shift, email, phone) in enumerate(TEAM_MEMBERS, 2):
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=2).value = shift
        ws.cell(row=row, column=3).value = email
        ws.cell(row=row, column=4).value = phone
    
    print(f"✓ Added {len(TEAM_MEMBERS)} team members")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    
    # Save workbook
    wb.save(EXCEL_FILE)
    wb.close()
    
    print(f"\n✅ Successfully saved {EXCEL_FILE}")
    print("\n" + "="*80)
    print("📊 TEAM MEMBERS SHEET CREATED")
    print("="*80)
    print(f"\nSheet2 contains {len(TEAM_MEMBERS)} team members:")
    for i, (name, shift, email, phone) in enumerate(TEAM_MEMBERS, 1):
        print(f"  {i:2d}. {name:20s} | {shift:8s} | {email:30s} | {phone}")
    
    print("\n✓ Next step: Update backend to read/write team members from Sheet2")
    
except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n")
