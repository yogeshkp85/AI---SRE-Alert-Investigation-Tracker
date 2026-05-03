#!/usr/bin/env python3
"""
Check what's actually in the Excel file
"""

import openpyxl

EXCEL_FILE = 'incident-tracker.xlsx'

print("\n" + "="*70)
print("EXCEL FILE DIAGNOSTIC")
print("="*70)

try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    print(f"\nSheet name: {ws.title}")
    print(f"Max row: {ws.max_row}")
    print(f"Max column: {ws.max_column}")
    
    print("\n" + "-"*70)
    print("HEADERS (Row 1):")
    print("-"*70)
    headers = []
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers.append(cell.value)
            print(f"  Col {idx}: {cell.value}")
    
    print("\n" + "-"*70)
    print("DATA ROWS:")
    print("-"*70)
    
    incident_count = 0
    empty_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        first_cell = row[0].value
        
        if first_cell is None or str(first_cell).strip() == '':
            empty_count += 1
            print(f"  Row {row_idx}: [EMPTY]")
        else:
            incident_count += 1
            # Get first few columns
            values = []
            for col_idx in range(1, min(6, ws.max_column + 1)):
                val = row[col_idx - 1].value
                if val:
                    values.append(str(val)[:20])
            print(f"  Row {row_idx}: {' | '.join(values)}")
    
    print("\n" + "-"*70)
    print("SUMMARY:")
    print("-"*70)
    print(f"Total rows (including header): {ws.max_row}")
    print(f"Data rows: {incident_count}")
    print(f"Empty rows: {empty_count}")
    print(f"Expected incidents: {incident_count}")
    
    wb.close()
    
except Exception as e:
    print(f"Error: {e}")

print("\n" + "="*70 + "\n")
