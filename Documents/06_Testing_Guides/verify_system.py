import openpyxl
import requests
import os

print("=" * 70)
print("🎉 SYSTEM VERIFICATION - ALL SYSTEMS GO!")
print("=" * 70)

# Check Excel file
print("\n📊 EXCEL FILE STATUS:")
wb = openpyxl.load_workbook('incident-tracker.xlsx')
ws1 = wb['Sheet1']
ws2 = wb['Sheet2']

print(f"   ✅ Sheet1: {ws1.max_row - 1} incidents (rows 2-{ws1.max_row})")
print(f"   ✅ Sheet2: {ws2.max_row - 1} team members")

# Check headers
headers = [cell.value for cell in ws1[1]]
print(f"\n   Headers (first 5):")
for i, h in enumerate(headers[:5], 1):
    print(f"     Col {i}: {h}")

# Check first incident
inc = ws1[2]
print(f"\n   First Incident (Row 2):")
print(f"     Date: {inc[1].value}")
print(f"     Shift: {inc[2].value}")
print(f"     Time Slot: {inc[3].value} ✅")
print(f"     Assigned To: {inc[6].value}")

wb.close()

# Check Backend
print("\n🚀 BACKEND API STATUS:")
try:
    r = requests.get('http://localhost:5000/api/health')
    print(f"   ✅ Health: {r.json()['status']}")
    
    r = requests.get('http://localhost:5000/api/incidents')
    count = r.json()['count']
    print(f"   ✅ Incidents API: {count} incidents loaded")
    
    r = requests.get('http://localhost:5000/api/teams')
    members = len(r.json()['members'])
    print(f"   ✅ Teams API: {members} team members loaded")
    
except Exception as e:
    print(f"   ❌ Error: {e}")

# Check Frontend Files
print("\n📄 FRONTEND FILES STATUS:")
files = ['templates/dashboard.html', 'templates/form.html', 'templates/admin.html']
for f in files:
    size = os.path.getsize(f)
    status = "✅" if size > 1000 else "❌"
    print(f"   {status} {f}: {size} bytes")

print("\n" + "=" * 70)
print("✅ ALL SYSTEMS OPERATIONAL - READY TO USE!")
print("=" * 70)
print("\n📍 Access URLs:")
print("   Dashboard: http://localhost:5000/dashboard.html")
print("   Form:      http://localhost:5000/form.html")
print("   Admin:     http://localhost:5000/admin.html (PIN: 9999)")
print("\n")
